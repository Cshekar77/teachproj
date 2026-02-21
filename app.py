from flask import Flask, render_template, request, redirect, abort, session, url_for, flash
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = "changeme_use_a_long_random_string"   # ← change in production
DB_PATH = "database.db"


# =====================================================
# ================= DB CONNECTION =====================
# =====================================================
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# =====================================================
# ================= AUTH DECORATORS ===================
# =====================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated


# =====================================================
# ===================== LOGIN / LOGOUT ================
# =====================================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        if session.get("role") == "admin":
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM Users WHERE username = ?", (username,)
        ).fetchone()
        conn.close()

        if user and check_password_hash(user["password"], password):
            session["user_id"]    = user["id"]
            session["username"]   = user["username"]
            session["role"]       = user["role"]
            session["faculty_id"] = user["faculty_id"]

            if user["role"] == "admin":
                return redirect(url_for("admin_dashboard"))
            else:
                return redirect(url_for("dashboard"))
        else:
            error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =====================================================
# ===================== DASHBOARD =====================
# =====================================================
@app.route("/")
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    selected_faculty = request.args.get("faculty_id", "")
    selected_day     = request.args.get("day_id", "")

    # Teachers can only see their own timetable
    if session.get("role") == "teacher":
        selected_faculty = session.get("faculty_id", "")

    query = """
        SELECT sm.id,
               f.name AS faculty_name,
               f.phone,
               f.email,
               s.subject_name,
               sec.section_name,
               sec.semester,
               d.day_name,
               ts.start_time,
               ts.end_time
        FROM SubjectMapping sm
        JOIN Faculty f ON sm.faculty_id = f.id
        JOIN Subject s ON sm.subject_id = s.id
        JOIN Section sec ON sm.section_id = sec.id
        LEFT JOIN Day d ON sm.day_id = d.id
        LEFT JOIN TimeSlot ts ON sm.timeslot_id = ts.id
        WHERE 1=1
    """
    params = []

    if selected_faculty:
        query += " AND sm.faculty_id = ?"
        params.append(selected_faculty)

    if selected_day:
        query += " AND sm.day_id = ?"
        params.append(selected_day)

    query += " ORDER BY d.id, ts.start_time"

    timetable    = cursor.execute(query, params).fetchall()
    faculty_list = cursor.execute("SELECT * FROM Faculty").fetchall()
    day_list     = cursor.execute("SELECT * FROM Day").fetchall()
    conn.close()

    return render_template(
        "index.html",
        timetable=timetable,
        faculty_list=faculty_list,
        day_list=day_list,
        selected_faculty=selected_faculty,
        selected_day=selected_day,
    )


# =====================================================
# ===================== ADMIN =========================
# =====================================================
@app.route("/admin")
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    data = {
        "faculty": cursor.execute("SELECT * FROM Faculty").fetchall(),
        "subject": cursor.execute("SELECT * FROM Subject").fetchall(),
        "section": cursor.execute("SELECT * FROM Section").fetchall(),
        "day":     cursor.execute("SELECT * FROM Day").fetchall(),
        "timeslot": cursor.execute("SELECT * FROM TimeSlot").fetchall(),
        "mapping": cursor.execute("""
            SELECT sm.id,
                   f.name AS faculty_name,
                   s.subject_name,
                   sec.section_name,
                   d.day_name,
                   ts.start_time,
                   ts.end_time
            FROM SubjectMapping sm
            JOIN Faculty f ON sm.faculty_id = f.id
            JOIN Subject s ON sm.subject_id = s.id
            JOIN Section sec ON sm.section_id = sec.id
            LEFT JOIN Day d ON sm.day_id = d.id
            LEFT JOIN TimeSlot ts ON sm.timeslot_id = ts.id
            ORDER BY sm.id
        """).fetchall(),
        "teacher_users": cursor.execute("""
            SELECT u.id, u.username, u.role, f.name AS faculty_name, f.id AS faculty_id
            FROM Users u
            LEFT JOIN Faculty f ON u.faculty_id = f.id
            WHERE u.role = 'teacher'
            ORDER BY u.id
        """).fetchall(),
    }

    conn.close()
    return render_template("admin_dashboard.html", **data)


# =====================================================
# ===================== DELETE ========================
# =====================================================
@app.route("/delete/<table>/<id>")
@admin_required
def delete_item(table, id):
    allowed_tables = [
        "Faculty", "Subject", "Section",
        "Day", "TimeSlot", "SubjectMapping"
    ]
    if table not in allowed_tables:
        abort(400, "Invalid table name")

    conn = get_db_connection()
    conn.execute(f"DELETE FROM {table} WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/admin")


# =====================================================
# ================= GENERIC ADD =======================
# =====================================================
def add_simple(table, fields, template):
    if request.method == "POST":
        conn = get_db_connection()
        values = [request.form.get(field) for field in fields]
        placeholders = ",".join(["?"] * len(fields))
        conn.execute(
            f"INSERT INTO {table} ({','.join(fields)}) VALUES ({placeholders})",
            values
        )
        conn.commit()
        conn.close()
        return redirect("/admin")
    return render_template(template)


# =====================================================
# ===================== ADD ROUTES ====================
# =====================================================
@app.route("/add/faculty", methods=["GET", "POST"])
@admin_required
def add_faculty():
    return add_simple("Faculty", ["id", "name", "phone", "email"], "add_faculty.html")

@app.route("/add/subject", methods=["GET", "POST"])
@admin_required
def add_subject():
    return add_simple("Subject", ["id", "subject_name", "subject_code"], "add_subject.html")

@app.route("/add/section", methods=["GET", "POST"])
@admin_required
def add_section():
    return add_simple("Section", ["id", "section_name", "semester"], "add_section.html")

@app.route("/add/day", methods=["GET", "POST"])
@admin_required
def add_day():
    return add_simple("Day", ["id", "day_name"], "add_day.html")

@app.route("/add/timeslot", methods=["GET", "POST"])
@admin_required
def add_timeslot():
    return add_simple("TimeSlot", ["id", "start_time", "end_time"], "add_timeslot.html")


# =====================================================
# ================= ADD SUBJECT MAPPING ===============
# =====================================================
@app.route("/add/subject_mapping", methods=["GET", "POST"])
@admin_required
def add_subject_mapping():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        cursor.execute("""
            INSERT INTO SubjectMapping (id, faculty_id, subject_id, section_id, day_id, timeslot_id)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            request.form["id"],
            request.form["faculty_id"],
            request.form["subject_id"],
            request.form["section_id"],
            request.form["day_id"],
            request.form["timeslot_id"],
        ))
        conn.commit()
        conn.close()
        return redirect("/admin")

    faculty  = cursor.execute("SELECT * FROM Faculty").fetchall()
    subject  = cursor.execute("SELECT * FROM Subject").fetchall()
    section  = cursor.execute("SELECT * FROM Section").fetchall()
    day      = cursor.execute("SELECT * FROM Day").fetchall()
    timeslot = cursor.execute("SELECT * FROM TimeSlot").fetchall()
    conn.close()

    return render_template("add_mapping.html",
        faculty=faculty, subject=subject, section=section,
        day=day, timeslot=timeslot)


# =====================================================
# ===================== EDIT ROUTES ===================
# =====================================================
@app.route("/edit/faculty/<id>", methods=["GET", "POST"])
@admin_required
def edit_faculty(id):
    conn = get_db_connection()
    if request.method == "POST":
        conn.execute("""
            UPDATE Faculty SET name=?, phone=?, email=? WHERE id=?
        """, (request.form["name"], request.form["phone"], request.form["email"], id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    faculty = conn.execute("SELECT * FROM Faculty WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("edit_faculty.html", faculty=faculty)


@app.route("/edit/subject/<id>", methods=["GET", "POST"])
@admin_required
def edit_subject(id):
    conn = get_db_connection()
    if request.method == "POST":
        conn.execute("""
            UPDATE Subject SET subject_name=?, subject_code=? WHERE id=?
        """, (request.form["subject_name"], request.form["subject_code"], id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    subject = conn.execute("SELECT * FROM Subject WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("edit_subject.html", subject=subject)


@app.route("/edit/section/<id>", methods=["GET", "POST"])
@admin_required
def edit_section(id):
    conn = get_db_connection()
    if request.method == "POST":
        conn.execute("""
            UPDATE Section SET section_name=?, semester=? WHERE id=?
        """, (request.form["section_name"], request.form["semester"], id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    section = conn.execute("SELECT * FROM Section WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("edit_section.html", section=section)


@app.route("/edit/day/<id>", methods=["GET", "POST"])
@admin_required
def edit_day(id):
    conn = get_db_connection()
    if request.method == "POST":
        conn.execute("UPDATE Day SET day_name=? WHERE id=?",
                     (request.form["day_name"], id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    day = conn.execute("SELECT * FROM Day WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("edit_day.html", day=day)


@app.route("/edit/timeslot/<id>", methods=["GET", "POST"])
@admin_required
def edit_timeslot(id):
    conn = get_db_connection()
    if request.method == "POST":
        conn.execute("""
            UPDATE TimeSlot SET start_time=?, end_time=? WHERE id=?
        """, (request.form["start_time"], request.form["end_time"], id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    timeslot = conn.execute("SELECT * FROM TimeSlot WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("edit_timeslot.html", timeslot=timeslot)


@app.route("/edit/mapping/<id>", methods=["GET", "POST"])
@admin_required
def edit_mapping(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == "POST":
        cursor.execute("""
            UPDATE SubjectMapping
            SET faculty_id=?, subject_id=?, section_id=?, day_id=?, timeslot_id=?
            WHERE id=?
        """, (
            request.form["faculty_id"], request.form["subject_id"],
            request.form["section_id"], request.form["day_id"],
            request.form["timeslot_id"], id
        ))
        conn.commit()
        conn.close()
        return redirect("/admin")

    mapping  = cursor.execute("SELECT * FROM SubjectMapping WHERE id=?", (id,)).fetchone()
    faculty  = cursor.execute("SELECT * FROM Faculty").fetchall()
    subject  = cursor.execute("SELECT * FROM Subject").fetchall()
    section  = cursor.execute("SELECT * FROM Section").fetchall()
    day      = cursor.execute("SELECT * FROM Day").fetchall()
    timeslot = cursor.execute("SELECT * FROM TimeSlot").fetchall()
    conn.close()

    return render_template("edit_mapping.html",
        mapping=mapping, faculty=faculty, subject=subject,
        section=section, day=day, timeslot=timeslot)


# =====================================================
# ================= ADD TEACHER USER ==================
# =====================================================
@app.route("/add/teacher_user", methods=["GET", "POST"])
@admin_required
def add_teacher_user():
    conn = get_db_connection()
    if request.method == "POST":
        username   = request.form["username"]
        password   = generate_password_hash(request.form["password"])
        faculty_id = request.form["faculty_id"]
        conn.execute("""
            INSERT INTO Users (username, password, role, faculty_id)
            VALUES (?, ?, 'teacher', ?)
        """, (username, password, faculty_id))
        conn.commit()
        conn.close()
        return redirect("/admin")
    faculty = conn.execute("SELECT * FROM Faculty").fetchall()
    conn.close()
    return render_template("add_teacher_user.html", faculty=faculty)




# =====================================================
# ================ CHANGE PASSWORD ====================
# =====================================================
@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    error   = None
    success = None

    if request.method == "POST":
        current  = request.form.get("current_password", "")
        new_pw   = request.form.get("new_password", "")
        confirm  = request.form.get("confirm_password", "")

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM Users WHERE id=?", (session["user_id"],)
        ).fetchone()

        if not check_password_hash(user["password"], current):
            error = "Current password is incorrect."
        elif len(new_pw) < 6:
            error = "New password must be at least 6 characters."
        elif new_pw != confirm:
            error = "New passwords do not match."
        else:
            conn.execute(
                "UPDATE Users SET password=? WHERE id=?",
                (generate_password_hash(new_pw), session["user_id"])
            )
            conn.commit()
            success = "Password updated successfully!"

        conn.close()

    return render_template("change_password.html",
        error=error, success=success, role=session.get("role"))



# =====================================================
# ============= EDIT / DELETE TEACHER USER ============
# =====================================================
@app.route("/edit/teacher_user/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_teacher_user(id):
    conn = get_db_connection()
    if request.method == "POST":
        username = request.form["username"]
        new_pw   = request.form.get("new_password", "").strip()
        if new_pw:
            pw_hash = generate_password_hash(new_pw)
            conn.execute("UPDATE Users SET username=?, password=? WHERE id=?",
                         (username, pw_hash, id))
        else:
            conn.execute("UPDATE Users SET username=? WHERE id=?", (username, id))
        conn.commit()
        conn.close()
        return redirect("/admin")

    user    = conn.execute("SELECT * FROM Users WHERE id=?", (id,)).fetchone()
    faculty = conn.execute("SELECT * FROM Faculty").fetchall()
    conn.close()
    return render_template("edit_teacher_user.html", user=user, faculty=faculty)


@app.route("/delete/teacher_user/<int:id>")
@admin_required
def delete_teacher_user(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM Users WHERE id=? AND role='teacher'", (id,))
    conn.commit()
    conn.close()
    return redirect("/admin")


# =====================================================
# ================= WEEKLY SUMMARY ====================
# =====================================================
@app.route("/weekly_summary")
@admin_required
def weekly_summary():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Faculty workload — classes per faculty sorted desc
    faculty_stats = cursor.execute("""
        SELECT f.name AS faculty_name, COUNT(sm.id) AS class_count
        FROM SubjectMapping sm
        JOIN Faculty f ON sm.faculty_id = f.id
        GROUP BY sm.faculty_id
        ORDER BY class_count DESC
    """).fetchall()

    # Classes per day
    day_stats = cursor.execute("""
        SELECT d.day_name, COUNT(sm.id) AS class_count
        FROM SubjectMapping sm
        JOIN Day d ON sm.day_id = d.id
        GROUP BY sm.day_id
        ORDER BY class_count DESC
    """).fetchall()

    # Full schedule
    schedule = cursor.execute("""
        SELECT f.name AS faculty_name,
               s.subject_name,
               sec.section_name,
               sec.semester,
               d.day_name,
               ts.start_time,
               ts.end_time
        FROM SubjectMapping sm
        JOIN Faculty f ON sm.faculty_id = f.id
        JOIN Subject s ON sm.subject_id = s.id
        JOIN Section sec ON sm.section_id = sec.id
        LEFT JOIN Day d ON sm.day_id = d.id
        LEFT JOIN TimeSlot ts ON sm.timeslot_id = ts.id
        ORDER BY f.name, d.id, ts.start_time
    """).fetchall()

    # Summary counts
    total_faculty  = cursor.execute("SELECT COUNT(*) FROM Faculty").fetchone()[0]
    total_classes  = cursor.execute("SELECT COUNT(*) FROM SubjectMapping").fetchone()[0]
    total_subjects = cursor.execute("SELECT COUNT(DISTINCT subject_id) FROM SubjectMapping").fetchone()[0]
    total_sections = cursor.execute("SELECT COUNT(DISTINCT section_id) FROM SubjectMapping").fetchone()[0]
    total_hours    = total_classes  # 1 class = 1 slot = assume 1 hr

    conn.close()
    return render_template("weekly_summary.html",
        faculty_stats=faculty_stats,
        day_stats=day_stats,
        schedule=schedule,
        total_faculty=total_faculty,
        total_classes=total_classes,
        total_hours=total_hours,
        total_subjects=total_subjects,
        total_sections=total_sections,
    )


# =====================================================
# ================= EXPORT TO EXCEL ===================
# =====================================================
@app.route("/export/weekly_summary")
@admin_required
def export_weekly_summary():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    conn = get_db_connection()
    cursor = conn.cursor()

    faculty_stats = cursor.execute("""
        SELECT f.name AS faculty_name, COUNT(sm.id) AS class_count
        FROM SubjectMapping sm
        JOIN Faculty f ON sm.faculty_id = f.id
        GROUP BY sm.faculty_id ORDER BY class_count DESC
    """).fetchall()

    day_stats = cursor.execute("""
        SELECT d.day_name, COUNT(sm.id) AS class_count
        FROM SubjectMapping sm
        JOIN Day d ON sm.day_id = d.id
        GROUP BY sm.day_id ORDER BY class_count DESC
    """).fetchall()

    schedule = cursor.execute("""
        SELECT f.name, s.subject_name, sec.section_name, sec.semester,
               d.day_name, ts.start_time, ts.end_time
        FROM SubjectMapping sm
        JOIN Faculty f ON sm.faculty_id = f.id
        JOIN Subject s ON sm.subject_id = s.id
        JOIN Section sec ON sm.section_id = sec.id
        LEFT JOIN Day d ON sm.day_id = d.id
        LEFT JOIN TimeSlot ts ON sm.timeslot_id = ts.id
        ORDER BY f.name, d.id, ts.start_time
    """).fetchall()
    conn.close()

    wb = Workbook()

    # ── Styles ──
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", start_color="4A6CF5")
    title_font = Font(name="Arial", bold=True, size=13, color="4A6CF5")
    body_font  = Font(name="Arial", size=10)
    center     = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="D5D9EC"),
        right=Side(style="thin", color="D5D9EC"),
        top=Side(style="thin", color="D5D9EC"),
        bottom=Side(style="thin", color="D5D9EC"),
    )
    alt_fill = PatternFill("solid", start_color="F4F6FB")

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = thin

    def style_data_row(ws, row, cols, alt=False):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            c.alignment = Alignment(vertical="center")
            c.border = thin
            if alt:
                c.fill = alt_fill

    # ── Sheet 1: Faculty Workload ──
    ws1 = wb.active
    ws1.title = "Faculty Workload"
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "Faculty Weekly Workload"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = center

    headers = ["Faculty Name", "Classes per Week", "Est. Hours"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=2, column=i, value=h)
    style_header_row(ws1, 2, 3)
    ws1.row_dimensions[2].height = 22

    for i, row in enumerate(faculty_stats, 3):
        ws1.cell(row=i, column=1, value=row["faculty_name"])
        ws1.cell(row=i, column=2, value=row["class_count"])
        ws1.cell(row=i, column=3, value=f"=B{i}")
        style_data_row(ws1, i, 3, alt=(i % 2 == 0))
        ws1.row_dimensions[i].height = 18

    total_row = len(faculty_stats) + 3
    ws1.cell(row=total_row, column=1, value="TOTAL")
    ws1.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)
    ws1.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row-1})")
    ws1.cell(row=total_row, column=2).font = Font(name="Arial", bold=True, size=10)
    ws1.cell(row=total_row, column=3, value=f"=SUM(C3:C{total_row-1})")
    ws1.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=10)
    style_data_row(ws1, total_row, 3)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 15

    # ── Sheet 2: Full Schedule ──
    ws2 = wb.create_sheet("Full Schedule")
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Complete Faculty Schedule"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center

    headers2 = ["Faculty", "Subject", "Section", "Semester", "Day", "Start Time", "End Time"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 7)
    ws2.row_dimensions[2].height = 22

    for i, row in enumerate(schedule, 3):
        vals = [row[0], row[1], row[2], f"Sem {row[3]}",
                row[4] or "—", row[5] or "—", row[6] or "—"]
        for j, v in enumerate(vals, 1):
            ws2.cell(row=i, column=j, value=v)
        style_data_row(ws2, i, 7, alt=(i % 2 == 0))
        ws2.row_dimensions[i].height = 18

    for col, width in zip("ABCDEFG", [26, 22, 16, 10, 12, 12, 12]):
        ws2.column_dimensions[col].width = width

    # ── Sheet 3: Classes per Day ──
    ws3 = wb.create_sheet("Classes per Day")
    ws3.row_dimensions[1].height = 30
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Classes per Day"
    ws3["A1"].font = title_font
    ws3["A1"].alignment = center

    for i, h in enumerate(["Day", "Class Count"], 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 2)

    for i, row in enumerate(day_stats, 3):
        ws3.cell(row=i, column=1, value=row["day_name"])
        ws3.cell(row=i, column=2, value=row["class_count"])
        style_data_row(ws3, i, 2, alt=(i % 2 == 0))
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 16

    # Save to buffer and send
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="weekly_summary.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ================= ERROR HANDLERS ====================
# =====================================================
@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403

# =====================================================
# ===================== RUN ===========================
# =====================================================
if __name__ == "__main__":
    app.run(debug=True)